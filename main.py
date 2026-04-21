"""
Script principal para ejecutar los cuatro algoritmos de NWJSSP
Genera archivos de resultados en formato Excel según especificaciones del curso

Algoritmos implementados:
1. Constructivo: Greedy determinístico
2. 2-opt: Búsqueda local, intercambio de trabajos consecutivos (best improvement)
3. Swap-10range: Búsqueda local, swap con distancia máxima 10 (best improvement)
4. Insertion-10range: Búsqueda local, inserción con distancia máxima 10 (first improvement)
"""

import os
import glob
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from constructive import ConstructiveAlgorithm
from two_opt import TwoOptSearch
from swap_10range import SwapRangeSearch
from insertion_10range import InsertionRangeSearch
from read_instances import read_nwjssp_instance

# ====== DIRECTORIO DE INSTANCIAS ======
INSTANCES_DIR = "instances"  # Directorio con archivos .txt de instancias


def create_results_workbook(algorithm_name):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def get_column_letter(col_num):
    col_letter = ""
    col = col_num
    while col >= 0:
        col_letter = chr(65 + (col % 26)) + col_letter
        col = col // 26 - 1
        if col < 0:
            break
    return col_letter


def add_results_sheet(workbook, instance_name, flow_time, computation_time, job_start_times):
    ws = workbook.create_sheet(instance_name)

    ws['A1'] = int(flow_time)
    ws['B1'] = int(round(computation_time))

    for idx, start_time in enumerate(job_start_times):
        col_letter = get_column_letter(idx)
        ws[f'{col_letter}2'] = int(start_time)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in [ws['A1'], ws['B1']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    for idx in range(len(job_start_times)):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = 12


# ====== FUNCIONES DE EJECUCIÓN POR ALGORITMO ======

def run_constructive_algorithm(n, m, operations, release_dates):
    algo = ConstructiveAlgorithm(n, m, operations, release_dates)
    return algo.solve()


def run_two_opt_algorithm(n, m, operations, release_dates):
    # Obtener solución inicial del constructivo
    constructive = ConstructiveAlgorithm(n, m, operations, release_dates)
    initial_solution, initial_flow_time, _ = constructive.solve()

    algo = TwoOptSearch(n, m, operations, release_dates)
    return algo.solve(initial_solution=initial_solution, initial_flow_time=initial_flow_time)


def run_swap_range_algorithm(n, m, operations, release_dates):
    constructive = ConstructiveAlgorithm(n, m, operations, release_dates)
    initial_solution, initial_flow_time, _ = constructive.solve()

    algo = SwapRangeSearch(n, m, operations, release_dates, max_range=10)
    return algo.solve(initial_solution=initial_solution, initial_flow_time=initial_flow_time)


def run_insertion_range_algorithm(n, m, operations, release_dates):
    constructive = ConstructiveAlgorithm(n, m, operations, release_dates)
    initial_solution, initial_flow_time, _ = constructive.solve()

    algo = InsertionRangeSearch(n, m, operations, release_dates, max_range=10)
    return algo.solve(initial_solution=initial_solution, initial_flow_time=initial_flow_time)


def process_instance(instance_file, algorithm_func):
    try:
        n, m, operations, release_dates, L = read_nwjssp_instance(instance_file)
        solution, flow_time, computation_time = algorithm_func(n, m, operations, release_dates)
        instance_name = os.path.splitext(os.path.basename(instance_file))[0]
        return instance_name, solution, flow_time, computation_time
    except Exception as e:
        print(f"  Error procesando {instance_file}: {str(e)}")
        return None


def run_algorithm_batch(instance_files, algorithm_func, algorithm_label, output_filename):
    """
    Ejecuta un algoritmo sobre todas las instancias y guarda resultados en Excel.

    Returns:
        total_time: tiempo total acumulado (ms)
        count: número de instancias procesadas
    """
    print(f"\nEjecutando {algorithm_label}...")
    print("-" * 70)

    wb = create_results_workbook(algorithm_label)
    total_time = 0
    count = 0

    for instance_file in instance_files:
        result = process_instance(instance_file, algorithm_func)
        if result:
            instance_name, solution, flow_time, computation_time = result
            print(f"  {instance_name:30s} | Z={flow_time:12.0f} | Tiempo={computation_time:8.2f}ms")
            total_time += computation_time
            count += 1
            add_results_sheet(wb, instance_name, flow_time, computation_time, solution)

    wb.save(output_filename)
    print(f"\n✓ Resultados guardados en {output_filename}")
    if count > 0:
        print(f"  Tiempo total: {total_time/1000:.2f}s, Promedio: {total_time/count:.2f}ms")

    return total_time, count


def main():
    print("=" * 70)
    print("NWJSSP - Solver: Constructivo, 2-opt, Swap-10, Insertion-10")
    print("=" * 70)

    instance_files = glob.glob(os.path.join(INSTANCES_DIR, "*.txt"))
    instance_files.sort()

    if not instance_files:
        print(f"Error: No se encontraron archivos de instancias en '{INSTANCES_DIR}'")
        return

    print(f"\nInstancias encontradas: {len(instance_files)}")

    # ====== ALGORITMOS Y SUS ARCHIVOS DE SALIDA ======
    algorithms = [
        (
            run_constructive_algorithm,
            "CONSTRUCTIVO (Greedy determinístico)",
            "NWJSSP_ArturoMurgueytio_Constructivo.xlsx",
        ),
        (
            run_two_opt_algorithm,
            "2-OPT (Swap consecutivos, Best Improvement)",
            "NWJSSP_ArturoMurgueytio_2opt.xlsx",
        ),
        (
            run_swap_range_algorithm,
            "SWAP-10 (Swap distancia ≤10, Best Improvement)",
            "NWJSSP_ArturoMurgueytio_Swap10.xlsx",
        ),
        (
            run_insertion_range_algorithm,
            "INSERTION-10 (Inserción distancia ≤10, First Improvement)",
            "NWJSSP_ArturoMurgueytio_Insertion10.xlsx",
        ),
    ]

    results_summary = []

    for algo_func, algo_label, output_file in algorithms:
        total_time, count = run_algorithm_batch(
            instance_files, algo_func, algo_label, output_file
        )
        results_summary.append((algo_label, count, total_time))

    # ====== RESUMEN COMPARATIVO ======
    print("\n" + "=" * 70)
    print("RESUMEN COMPARATIVO")
    print("=" * 70)
    print(f"{'Algoritmo':<45} | {'Inst.':<6} | {'Total':>10} | {'Promedio':>10}")
    print("-" * 70)
    for label, count, total in results_summary:
        avg = total / count if count > 0 else 0
        print(f"  {label:<43} | {count:<6} | {total/1000:>7.2f}s  | {avg:>8.2f}ms")

    print("\n✓ Ejecución completada exitosamente")
    print("\nArchivos generados:")
    for _, _, output_file in algorithms:
        print(f"  - {output_file}")
    print()


if __name__ == "__main__":
    main()